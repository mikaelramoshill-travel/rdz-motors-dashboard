"""
Car Factory Production Data Generator
-------------------------------------
Automates creation of a formatted Excel workbook (factory_data.xlsx) with
Daily, Monthly and Yearly production data, AND exports the same data to
`data.js` so the local dashboard works instantly by just opening the HTML file.

Run:  python generate_data.py

Dependencies: openpyxl  (pip install openpyxl)
"""

from __future__ import annotations

import json
import random
from datetime import date, datetime, timedelta
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    HAVE_OPENPYXL = True
except ImportError:  # pragma: no cover - graceful fallback
    HAVE_OPENPYXL = False


HERE = Path(__file__).resolve().parent

MODELS = ["Sedan", "SUV", "Electric", "Truck"]
LINES = ["Line A", "Line B", "Line C", "Line D"]

# Deterministic-ish randomness so the numbers look stable between runs
random.seed(42)


# --------------------------------------------------------------------------- #
#  Data generation
# --------------------------------------------------------------------------- #
def _model_split(total: int) -> dict[str, int]:
    """Split a total unit count across the car models."""
    weights = [0.34, 0.30, 0.22, 0.14]  # Sedan, SUV, EV, Truck
    parts = [int(total * w) for w in weights]
    parts[0] += total - sum(parts)  # keep the sum exact
    return dict(zip(MODELS, parts))


def build_daily() -> list[dict]:
    """Hour-by-hour production for the current shift (today)."""
    rows = []
    start = 6  # 06:00 factory start
    for i in range(12):  # 12 working hours
        hour = f"{start + i:02d}:00"
        target = 45
        produced = target - random.randint(-6, 8)
        defects = max(0, round(produced * random.uniform(0.01, 0.05)))
        downtime = random.choice([0, 0, 0, 5, 8, 12])
        rows.append(
            {
                "period": hour,
                "produced": produced,
                "target": target,
                "defects": defects,
                "downtime": downtime,
                "efficiency": round(produced / target * 100, 1),
                **_model_split(produced),
            }
        )
    return rows


def build_monthly() -> list[dict]:
    """Day-by-day production for the last 30 days."""
    rows = []
    today = date.today()
    for d in range(29, -1, -1):
        day = today - timedelta(days=d)
        weekend = day.weekday() >= 5
        target = 380 if not weekend else 180
        produced = target - random.randint(-40, 55)
        defects = max(0, round(produced * random.uniform(0.015, 0.045)))
        downtime = random.randint(0, 45)
        rows.append(
            {
                "period": day.strftime("%b %d"),
                "produced": produced,
                "target": target,
                "defects": defects,
                "downtime": downtime,
                "efficiency": round(produced / target * 100, 1),
                **_model_split(produced),
            }
        )
    return rows


def build_yearly() -> list[dict]:
    """Month-by-month production for the last 12 months."""
    rows = []
    today = date.today().replace(day=1)
    months = []
    cursor = today
    for _ in range(12):
        months.append(cursor)
        # step back one month
        prev = cursor.replace(day=1) - timedelta(days=1)
        cursor = prev.replace(day=1)
    months.reverse()

    for m in months:
        target = 9200
        produced = target - random.randint(-1200, 1400)
        defects = max(0, round(produced * random.uniform(0.02, 0.04)))
        downtime = random.randint(120, 600)
        rows.append(
            {
                "period": m.strftime("%b %Y"),
                "produced": produced,
                "target": target,
                "defects": defects,
                "downtime": downtime,
                "efficiency": round(produced / target * 100, 1),
                **_model_split(produced),
            }
        )
    return rows


# --------------------------------------------------------------------------- #
#  Excel export
# --------------------------------------------------------------------------- #
HEADER = ["Period", "Produced", "Target", "Defects", "Downtime (min)",
          "Efficiency %", *MODELS]
FIELDS = ["period", "produced", "target", "defects", "downtime",
          "efficiency", *MODELS]

HEADER_FILL = PatternFill("solid", fgColor="2E1065")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, color="6D28D9", size=16)
THIN = Side(style="thin", color="E5E1F0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")


def _write_sheet(ws, title: str, rows: list[dict], chart_type: str):
    ws.sheet_view.showGridLines = False

    ws["A1"] = f"RDZ MOTORS  —  {title}"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADER))
    ws.row_dimensions[1].height = 26

    header_row = 3
    for c, name in enumerate(HEADER, start=1):
        cell = ws.cell(row=header_row, column=c, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER

    for r, row in enumerate(rows, start=header_row + 1):
        for c, field in enumerate(FIELDS, start=1):
            cell = ws.cell(row=r, column=c, value=row[field])
            cell.alignment = CENTER
            cell.border = BORDER
            if field == "efficiency":
                cell.number_format = '0.0"%"'
                if row[field] >= 100:
                    cell.fill = PatternFill("solid", fgColor="DCFCE7")
                elif row[field] >= 90:
                    cell.fill = PatternFill("solid", fgColor="FEF9C3")
                else:
                    cell.fill = PatternFill("solid", fgColor="FEE2E2")

    # Totals row
    total_row = header_row + 1 + len(rows)
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    for c in range(2, len(HEADER) + 1):
        col = get_column_letter(c)
        top = header_row + 1
        bottom = total_row - 1
        if HEADER[c - 1] == "Efficiency %":
            ws.cell(row=total_row, column=c,
                    value=f"=AVERAGE({col}{top}:{col}{bottom})").number_format = '0.0"%"'
        else:
            ws.cell(row=total_row, column=c, value=f"=SUM({col}{top}:{col}{bottom})")
        ws.cell(row=total_row, column=c).font = Font(bold=True)
        ws.cell(row=total_row, column=c).alignment = CENTER

    # Column widths
    ws.column_dimensions["A"].width = 14
    for c in range(2, len(HEADER) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 13

    # Chart
    data_min = header_row + 1
    data_max = header_row + len(rows)
    if chart_type == "line":
        chart = LineChart()
    else:
        chart = BarChart()
        chart.type = "col"
    chart.title = f"{title} — Produced vs Target"
    chart.height = 8
    chart.width = 20
    data = Reference(ws, min_col=2, max_col=3, min_row=header_row, max_row=data_max)
    cats = Reference(ws, min_col=1, min_row=data_min, max_row=data_max)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, f"A{total_row + 2}")


def export_excel(daily, monthly, yearly) -> Path:
    wb = Workbook()
    wb.remove(wb.active)

    _write_sheet(wb.create_sheet("Daily"), "Daily Production (Today)", daily, "bar")
    _write_sheet(wb.create_sheet("Monthly"), "Monthly Production (30 Days)", monthly, "line")
    _write_sheet(wb.create_sheet("Yearly"), "Yearly Production (12 Months)", yearly, "bar")

    out = HERE / "factory_data.xlsx"
    wb.save(out)
    return out


# --------------------------------------------------------------------------- #
#  JS export (so the dashboard works with a double-click, no server needed)
# --------------------------------------------------------------------------- #
def export_js(daily, monthly, yearly) -> Path:
    payload = {
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "models": MODELS,
        "daily": daily,
        "monthly": monthly,
        "yearly": yearly,
    }
    out = HERE / "data.js"
    out.write_text(
        "// Auto-generated by generate_data.py — do not edit by hand.\n"
        "window.FACTORY_DATA = " + json.dumps(payload, indent=2) + ";\n",
        encoding="utf-8",
    )
    return out


# --------------------------------------------------------------------------- #
def main() -> None:
    daily, monthly, yearly = build_daily(), build_monthly(), build_yearly()

    js_path = export_js(daily, monthly, yearly)
    print(f"✅ Dashboard data written: {js_path.name}")

    if HAVE_OPENPYXL:
        xlsx_path = export_excel(daily, monthly, yearly)
        print(f"✅ Excel workbook written:  {xlsx_path.name}")
    else:
        print("⚠️  openpyxl not installed — skipped Excel export.")
        print("    Install it with:  pip install openpyxl")

    print("\nOpen dashboard.html in your browser to view the dashboard.")


if __name__ == "__main__":
    main()
